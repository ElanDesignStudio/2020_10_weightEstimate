#------ Versions:
# 001   basic file reader
# 002   better layout, limits etc
# 003   annotation interesting points

import openpyxl
import io
import numpy as np

import matplotlib
import matplotlib.pyplot as plt
from   matplotlib.ticker import (AutoMinorLocator, MultipleLocator)


# 1 ---- define a class to store weight information ----------------------------------------
class weightItem:
    def __init__(self):
        self.weight = 0
        self.cg = [0,0,0]
        self.group = ""
        self.subgroup = ""
        self.item = ""


# 2 ---- open spreadsheet ---------------------------------------------------------------
xlsx_filename = "C:\\Users\\Public\\code_example\\weights.xlsx"         #   a variable for holding the name of our spreadsheet

with open(xlsx_filename, "rb") as f:                                    #   Open the file from disk - will read the last SAVED version
    in_mem_file = io.BytesIO(f.read())                                  #   read the file into RAM - means we can read the file even if it's open in Excel

spreadSheet = openpyxl.load_workbook(in_mem_file,read_only=True,keep_vba=True,data_only=True)   #   load the workbook
weightWorksheet = spreadSheet['Weights']                                                        #   get the 'Weights' spreadsheet
                                                                                                #   MUST be exactly the same as the sheet name - capitals and all

# 3 ---- read the data _____________________________________________________________________
allTheWeights = []          # where we're storing the data we read from the spreadsheet
rowInSpreadsheet=2          # start reading from row 2

while True:
    wi = weightItem()       #   createan 'instance' of our class to store the weight info
    weight  = weightWorksheet.cell(row=rowInSpreadsheet, column=8).value            #   read the cell contents
    group   = weightWorksheet.cell(row=rowInSpreadsheet, column=1).value            #   read the cell contents

    if(weight is not None) and (weight > 0):                                                                 #   check - did we read any weight data?

        wi.weight = weight

        subgroup= weightWorksheet.cell(row=rowInSpreadsheet, column=2).value        #   read the cell contents
        item    = weightWorksheet.cell(row=rowInSpreadsheet, column=3).value        #

        wi.cg[0] = weightWorksheet.cell(row=rowInSpreadsheet, column=9).value       #   store LCG
        wi.cg[1] = weightWorksheet.cell(row=rowInSpreadsheet, column=10).value      #   store TCG
        wi.cg[2] = weightWorksheet.cell(row=rowInSpreadsheet, column=11).value      #   store VCG

        if group != None:               #   if we read a value for 'group'....
            wi.group = group            #   ...store it

        if subgroup != None:            #   if we read a value for 'subgroup'....
            wi.subgroup = subgroup      #   ...store it
    
        if item != None:                #   if we read a value for 'item'....
            wi.item = item              #   ...store it

        allTheWeights.append(wi)        #   add the data we read into our list of weight items                                        

    rowInSpreadsheet+=1                 #   move on to reading the next row

    if( rowInSpreadsheet > 10000 ):     #   know when to stop
        break

    if group is not None and group=='[end]':
        break

# 4 ---- okay, now we've read all the data, let's draw pictures.

# 5 ---- plot configuration ----------------------------------------------------------
fig=plt.figure(figsize=(12,9))
ax=fig.add_subplot(1,1,1)
plt.style.use('seaborn-whitegrid')
cmapLines=plt.cm.get_cmap('nipy_spectral')  #   get a 'colormap' - handy for specifying colours!


# 6 ---- add the data ------------
weight = []
posLong = []
posTrans = []
posVert = []

for item in allTheWeights:
    weight.append(   item.weight )
    posLong.append(  item.cg[0] )
    posTrans.append( item.cg[1] )
    posVert.append(  item.cg[2] )
    
scatter = plt.scatter(posLong,posVert,s=weight,c=weight,label="Long/vert",cmap='nipy_spectral')  # see: https://matplotlib.org/3.1.1/gallery/shapes_and_collections/scatter.html#sphx-glr-gallery-shapes-and-collections-scatter-py

# 7 ---- add a legend with the unique colors from the scatter
legend1 = ax.legend(*scatter.legend_elements(),loc="upper left", title="Weights",framealpha=0.5,frameon=True)      # see: https://matplotlib.org/3.1.1/gallery/lines_bars_and_markers/scatter_with_legend.html
ax.add_artist(legend1)


# 8 ---- annotate interesting items --------------------------------------------
for item in allTheWeights:
    if item.weight > 1000:
        col = cmapLines(0.5)
        arrowEndX = item.cg[0]
        arrowEndY = item.cg[2]
        tag = item.group + ": " + item.subgroup + ": " + item.item
        ax.annotate(tag,xy=(arrowEndX,arrowEndY),xycoords='data',xytext=(-150,30),textcoords='offset points',size=9,bbox=dict(boxstyle="square",fc=col,ec=col,alpha=0.2),arrowprops=dict(arrowstyle="->",ec=col))


# 9 ---- plot tidy up ----------------------------------------------------------
ax.set_xlim(0,10)
ax.set_ylim(0,2)
ax.set_title('Weight estimate',fontsize=12)
ax.set_xlabel('Longitudinal position')
ax.set_ylabel('Vertical position')


# 10 ---- Change tick markers
ax.xaxis.set_major_locator(MultipleLocator(1))
ax.xaxis.set_minor_locator(MultipleLocator(0.2))
ax.yaxis.set_major_locator(MultipleLocator(0.2))
ax.yaxis.set_minor_locator(MultipleLocator(0.1))
ax.grid(which='major',color='#CCCCCC',linestyle='-')        # draw the grid
ax.grid(which='minor',color='#CCCCCC',linestyle=':')

# 11 ---- add a colour bar to show the scale
plt.colorbar()

# 12 ---- show the plot
plt.show()


